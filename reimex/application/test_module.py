from reimex.models import SoenModel
from actualSpot.models import Aspot
from openpyxl import Workbook, load_workbook
import datetime
import os

now=datetime.datetime.now()

def func():                 #member_choice.htmlから呼ばれてる
    selected_spot=Aspot.objects.filter(checking=True)
    spotName=[]
    superVisor=[]
    firstSubcontractor=[]
    X=[]
    XSubcontractor=[]
    
    members_name=[]
    members_code=[]
    members_spotName=[]
    members_kanaName=[]
    members_age=[]
    members_djgroup=[]
    members_djgroup2=[]
    members_occupation=[]
    members_refer=[]
    members_birthday=[]
    members_dateOfEmployment=[]
    members_workingFirstDate=[]
    members_yearsOfExperience=[]
    members_postalCode=[]
    members_address=[]
    members_telephoneNumber=[]
    members_familyAdd=[]
    members_tel=[]
    members_medicalCheckUp=[]
    members_bloodPressure=[]
    members_bloodType=[]
    members_specialHealthCheck=[]
    members_SHCkind=[]
    members_healthInsurance=[]
    members_HIlast4digits=[]
    members_pensionInsurance=[]
    members_employmentinsurance=[]
    members_EIlast4digits=[]
    members_retirementCooperation=[]
    members_specialEducation=[]
    members_skillTraning=[]
    members_licence=[]
    members_dateOfAdmission=[]
    members_file1=[]
    members_file2=[]
    members_file3=[]
    members_file4=[]
    members_file5=[]
    members_file6=[]
    members_file7=[]
    members_file8=[]
    members_file9=[]
    members_file10=[]

# ***************************************************************
    for selected in selected_spot:
        spotName.append(selected.spotName)
        superVisor.append(selected.superVisor)
        firstSubcontractor.append(selected.firstSubcontractor)
        X.append(selected.X)
        XSubcontractor.append(selected.XSubcontractor)
        
    for member in selected.members.all():
        members_name.append(member.name)
        members_code.append(member.code)
        members_kanaName.append(member.kanaName)
        members_age.append(member.age)
        members_djgroup.append(member.djgroup)
        members_djgroup2.append(member.djgroup2)
        members_occupation.append(member.occupation)
        members_refer.append(member.refer)
        members_birthday.append(member.birthday)
        members_dateOfEmployment.append(member.dateOfEmployment)
        members_workingFirstDate.append(member.workingFirstDate)
        members_yearsOfExperience.append(member.yearsOfExperience)
        members_postalCode.append(member.postalCode)
        members_address.append(member.address)
        members_telephoneNumber.append(member.telephoneNumber)
        members_familyAdd.append(member.familyAdd)
        members_tel.append(member.tel)
        members_medicalCheckUp.append(member.medicalCheckUp)
        members_bloodPressure.append(member.bloodPressure)
        members_bloodType.append(member.bloodType)
        members_specialHealthCheck.append(member.specialHealthCheck)
        members_SHCkind.append(member.SHCkind)
        members_healthInsurance.append(member.healthInsurance)
        members_HIlast4digits.append(member.HIlast4digits)
        members_pensionInsurance.append(member.pensionInsurance)
        members_employmentinsurance.append(member.employmentinsurance)
        members_EIlast4digits.append(member.EIlast4digits)
        members_retirementCooperation.append(member.retirementCooperation)
        members_specialEducation.append(member.specialEducation)
        members_skillTraning.append(member.skillTraning)
        members_licence.append(member.licence)
        members_dateOfAdmission.append(member.dateOfAdmission)
        members_file1.append(member.file1)
        members_file2.append(member.file2)
        members_file3.append(member.file3)
        members_file4.append(member.file4)
        members_file5.append(member.file5)
        members_file6.append(member.file6)
        members_file7.append(member.file7)
        members_file8.append(member.file8)
        members_file9.append(member.file9)
        members_file10.append(member.file10)
        
#.append(member.) ***************************************************************
    wb= load_workbook(filename=r"C:\Users\PCUSER\Desktop\tesdata\input_data\soen_member_export0.xlsm",read_only=False,keep_vba=True)
    ws=wb['TranscriptionData']
    ws.cell(3,3,value=spotName[0])
    ws.cell(4,3,value=superVisor[0])
    ws.cell(5,3,value=firstSubcontractor[0])
    ws.cell(6,3,value=X[0])
    ws.cell(7,3,value=XSubcontractor[0])
    for i in range(0,len(members_name)):
        ws.cell(i+10,3,value=members_name[i])
        ws.cell(i+10,2,value=members_code[i])
        ws.cell(i+10,4,value=members_kanaName[i])
        ws.cell(i+10,5,value=members_age[i])      
        ws.cell(i+10,6,value=members_djgroup[i])
        ws.cell(i+10,7,value=members_djgroup2[i])
        ws.cell(i+10,8,value=members_occupation[i])
        ws.cell(i+10,9,value=members_refer[i])
        ws.cell(i+10,10,value=members_birthday[i])
        ws.cell(i+10,11,value=members_dateOfEmployment[i])
        ws.cell(i+10,12,value=members_workingFirstDate[i])
        ws.cell(i+10,13,value=members_yearsOfExperience[i])
        ws.cell(i+10,14,value=members_postalCode[i])
        ws.cell(i+10,15,value=members_address[i])
        ws.cell(i+10,16,value=members_telephoneNumber[i])
        ws.cell(i+10,17,value=members_familyAdd[i])
        ws.cell(i+10,18,value=members_tel[i])
        ws.cell(i+10,19,value=members_medicalCheckUp[i])
        ws.cell(i+10,20,value=members_bloodPressure[i])
        ws.cell(i+10,21,value=members_bloodType[i])
        ws.cell(i+10,22,value=members_specialHealthCheck[i])
        ws.cell(i+10,23,value=members_SHCkind[i])
        ws.cell(i+10,24,value=members_healthInsurance[i])
        ws.cell(i+10,25,value=members_HIlast4digits[i])
        ws.cell(i+10,26,value=members_pensionInsurance[i])
        ws.cell(i+10,27,value=members_employmentinsurance[i])
        ws.cell(i+10,28,value=members_EIlast4digits[i])
        ws.cell(i+10,29,value=members_retirementCooperation[i])
        ws.cell(i+10,30,value=members_specialEducation[i])
        ws.cell(i+10,31,value=members_skillTraning[i])
        ws.cell(i+10,32,value=members_licence[i])
        ws.cell(i+10,33,value=members_dateOfAdmission[i])
        ws.cell(i+10,35,value=str(members_file1[i]))
        ws.cell(i+10,36,value=str(members_file2[i]))
        ws.cell(i+10,37,value=str(members_file3[i]))
        ws.cell(i+10,38,value=str(members_file4[i]))
        ws.cell(i+10,39,value=str(members_file5[i]))
        ws.cell(i+10,40,value=str(members_file6[i]))
        ws.cell(i+10,41,value=str(members_file7[i]))
        ws.cell(i+10,42,value=str(members_file8[i]))
        ws.cell(i+10,43,value=str(members_file9[i]))
        ws.cell(i+10,44,value=str(members_file10[i]))


    wb.save(r"C:\Users\PCUSER\Desktop\tesdata\output_data\{0:%Y%m%d_%H%M%S}.xlsm".format(now))




if __name__=="__main__":
    func()
    