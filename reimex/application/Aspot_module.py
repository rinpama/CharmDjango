from reimex.models import SoenModel,SpecialEducationModel,SkillTraningModel,LicenceModel
from actualSpot.models import Aspot,mostRecent,gContractor,Vehicle
from openpyxl import Workbook, load_workbook
import datetime
import os

now=datetime.datetime.now()

def func(self,*args,**kwargs):
    selected_spot=Aspot.objects.filter(checking=True)
    gContractor_all=[] #元請会社
    X_all=[]
    X_vehicle_all=[]
    Xx_all=[]
    X0={}
    Xx_lic_all=[]#*******
    for selected in selected_spot:
        spotName=selected.spotName #現場名称
        superVisor=selected.superVisor #現場所長名
        gContractor_all.append(selected.gContractor.all()) #元請会社all
        X_all.append(selected.XSubcontractor.all()) #基本装苑all
        Xx_all.append(selected.XxSubcontractor.all()) #基本職方all
      
            
            
    #元請会社 ************************************************************************************************
    for g in gContractor_all:
        g.gName=g[0].gName
        g.postalCode=g[0].postalCode
        g.address=g[0].address
        g.telephoneNumber=g[0].telephoneNumber
        g.FaxNumber=g[0].FaxNumber     
    # print(g.telephoneNumber)

    #基本装苑all *********************************************************************************************
    xrecentName=[]
    xpostalCode=[]
    xaddress=[]
    xtelephoneNumber=[]
    xFaxNumber=[]
    xrepresentative=[]
    xsaftyManager=[]
    xsaftyPromoter=[]
    xleadEngineer=[]
    xfile0=[]
    xInsurance_station=[]
    xInsurance_station_number=[]
    xhealthInsurance=[]
    xfile1=[]
    xpensionInsurance=[]
    xfile2=[]
    xemploymentinsurance=[]
    xemploymentinsurance_number=[]
    xfile3=[]
    xretirementCooperation=[]
    xfile4=[]
    xconstractionPermit=[]
    xconstractionPermit_chuji=[]
    xconstractionPermit_ippan=[]
    xconstractionPermit_nendo=[]
    xconstractionPermit_num=[]
    xconstractionPermit_gat=[]
    xfile5=[]
    xcontractDay=[]
    xfirstActionDay=[]
    xfinishActionDay=[]
    xcontractionDetail=[]
    xNo_1_specific_skill_foreigner=[]
    xforeignWorker=[]
    xforeignIntern=[]
    xcarryonMachine=[]
    xvehicleNumber=[]
    xmrecent=[]
    xvehicleNumber=[]
    xfirstUse=[]
    xfinishUse=[]
    xvehicleModel=[]
    xfirstInspection=[]
    xfinishInspection=[]
    xfile1=[]
    xliabilityComp=[]
    xliability_num=[]
    xfirstliability=[]
    xfinishliability=[]
    xfile2=[]
    xvoluntaryInsureComp=[]
    xvoluntaryInsureNum=[]
    xfirstVoluntary=[]
    xfinishVoluntary=[]
    xvoluntary_object=[]
    xvoluntary_human=[]
    xvoluntary_passenger=[]
    xfile3=[]
    dvName=[]
    xdrive_departure=[]
    xdrive_waypoint1=[]
    xdrive_waypoint2=[]
    xdrive_arrival=[]
    xmobileCrane=[]
    xdangerous=[]
    xmemcode=[]
    xmemdjgroup=[]
    xmemname=[]
    xmemkanaName=[]
    xmemdjgroup2=[]
    xmemoccupation=[]
    xmemrefer=[]
    xmembirthday=[]
    xmemage=[]
    xmemdateOfEmployment=[]
    xmemworkingFirstDate=[]
    xmemyearsOfExperience=[]
    xmempostalCode=[]
    xmemaddress=[]
    xmemtelephoneNumber=[]
    xmemfamilyAdd=[]
    xmemtel=[]
    xmemmedicalCheckUp=[]
    xmembloodPressure=[]
    xmembloodType=[]
    xmemspecialHealthCheck=[]
    xmemSHCkind=[]
    xmemsaftyManager=[]
    xmemsaftyPromoter=[]
    xmemleadEngineer=[]
    xmemfile0=[]
    xmemInsurance_station=[]
    xmemInsurance_station_number=[]
    xmemhealthInsurance=[]
    xmemHIlast4digits=[]
    xmemfile1=[]
    xmempensionInsurance=[]
    xmemfile2=[]
    xmememploymentinsurance=[]
    xmemEIlast4digits=[]
    xmemfile3=[]
    xmemretirementCooperation=[]
    xmemfile4=[]
    xmemspecialEducation=[]
    xmemfile5=[]
    xmemskillTraning=[]
    xmemfile6=[]
    xmemlicence=[]
    xmemfile7=[]
    xmemdriverLicence_number=[]
    xmemfile8=[]
    xmemdateOfAdmission=[]
    xmemfile9=[]
    xmemfile10=[]
    xmemconstractionPermit=[]
    xmemconstractionPermit_chuji=[]
    xmemconstractionPermit_ippan=[]
    xmemconstractionPermit_nendo=[]
    xmemconstractionPermit_num=[]
    xmemconstractionPermit_gat=[]
    xmemfile5=[]
    xmemcontractDay=[]
    xmemfirstActionDay=[]
    xmemfinishActionDay=[]
    xmemcontractionDetail=[]
    xmemNo_1_specific_skill_foreigner=[]
    xmemforeignWorker=[]
    xmemforeignIntern=[]
    xmemcarryonMachine=[]
    xmemmvehicle=[]
    xmemmobileCrane=[]
    xmemdangerous=[]




    for x1 in X_all:
        for x in x1.all():
            xrecentName.append(x.recentName)
            xpostalCode.append(x.postalCode)
            xaddress.append(x.address)
            xtelephoneNumber.append(x.telephoneNumber)
            xFaxNumber.append(x.FaxNumber)
            xrepresentative.append(x.representative)
            xsaftyManager.append(x.saftyManager)
            xsaftyPromoter.append(x.saftyPromoter)
            xleadEngineer.append(x.leadEngineer)
            xfile0.append(x.file0)
            xInsurance_station.append(x.Insurance_station)
            xInsurance_station_number.append(x.Insurance_station_number)
            xhealthInsurance.append(x.healthInsurance)
            xfile1.append(x.file1)
            xpensionInsurance.append(x.pensionInsurance)
            xfile2.append(x.file2)
            xemploymentinsurance.append(x.employmentinsurance)
            xemploymentinsurance_number.append(x.employmentinsurance_number)
            xfile3.append(x.file3)
            xretirementCooperation.append(x.retirementCooperation)
            xfile4.append(x.file4)
            xconstractionPermit.append(x.constractionPermit)
            xconstractionPermit_chuji.append(x.constractionPermit_chuji)
            xconstractionPermit_ippan.append(x.constractionPermit_ippan)
            xconstractionPermit_nendo.append(x.constractionPermit_nendo)
            xconstractionPermit_num.append(x.constractionPermit_num)
            xconstractionPermit_gat.append(x.constractionPermit_gat)
            xfile5.append(x.file5)
            xcontractDay.append(x.contractDay)
            xfirstActionDay.append(x.firstActionDay)
            xfinishActionDay.append(x.finishActionDay)
            xcontractionDetail.append(x.contractionDetail)
            xNo_1_specific_skill_foreigner.append(x.No_1_specific_skill_foreigner)
            xforeignWorker.append(x.foreignWorker)
            xforeignIntern.append(x.foreignIntern)
            xcarryonMachine.append(x.carryonMachine)
            
            # for sp in Xx.specialEducation.all():
            #         xxspkey.append(Xx.name)
            #         xxspvalue.append(sp.name)
            #         for m,n in zip(xxspkey,xxspvalue):
            #             l=(m,n)
            #             xxsp.append(l)
            #             dxxsp=set(xxsp)
            #     xxfile5.append(Xx.file5)
            
            for veh in x.vehicle.all():#複数選択可はリストORタプル等で取得して処理する
                    xvehicleNumber.append(veh.vehicleNumber)
                    xmrecent.append(veh.mrecent)
                    xvehicleNumber.append(veh.vehicleNumber[0])
                    xfirstUse.append(veh.firstUse)
                    xfinishUse.append(veh.finishUse)
                    xvehicleModel.append(veh.vehicleModel)
                    xfirstInspection.append(veh.firstInspection)
                    xfinishInspection.append(veh.finishInspection)
                    xfile1.append(veh.file1)
                    xliabilityComp.append(veh.liabilityComp)
                    xliability_num.append(veh.liability_num)
                    xfirstliability.append(veh.firstliability)
                    xfinishliability.append(veh.finishliability)
                    xfile2.append(veh.file2)
                    xvoluntaryInsureComp.append(veh.voluntaryInsureComp)
                    xvoluntaryInsureNum.append(veh.voluntaryInsureNum)
                    xfirstVoluntary.append(veh.firstVoluntary)
                    xfinishVoluntary.append(veh.finishVoluntary)
                    xvoluntary_object.append(veh.voluntary_object)
                    xvoluntary_human.append(veh.voluntary_human)
                    xvoluntary_passenger.append(veh.voluntary_passenger)
                    xfile3.append(veh.file3)
                    for dv in veh.driver.all():
                        dvName.append(dv.name)
                    xdrive_departure .append(veh.drive_departure)
                    xdrive_waypoint1 .append(veh.drive_waypoint1)
                    xdrive_waypoint2 .append(veh.drive_waypoint2)
                    xdrive_arrival .append(veh.drive_arrival)
            xmobileCrane.append(x.mobileCrane)
            xdangerous.append(x.dangerous)
            for mem in x.members.all():
                xmemcode.append(mem.code)
                xmemdjgroup.append(mem.djgroup)
                xmemname.append(mem.name)
                xmemkanaName.append(mem.kanaName)
                xmemdjgroup2.append(mem.djgroup2)
                xmemoccupation.append(mem.occupation)
                xmemrefer.append(mem.refer)
                xmembirthday.append(mem.birthday)
                xmemage.append(mem.age)
                xmemdateOfEmployment.append(mem.dateOfEmployment)
                xmemworkingFirstDate.append(mem.workingFirstDate)
                xmemyearsOfExperience.append(mem.yearsOfExperience)
                xmempostalCode.append(mem.postalCode)
                xmemaddress.append(mem.address)
                xmemtelephoneNumber.append(mem.telephoneNumber)
                xmemfamilyAdd.append(mem.familyAdd)
                xmemtel.append(mem.tel)
                xmemmedicalCheckUp.append(mem.medicalCheckUp)
                xmembloodPressure.append(mem.bloodPressure)
                xmembloodType.append(mem.bloodType)
                xmemspecialHealthCheck.append(mem.specialHealthCheck)
                xmemSHCkind.append(mem.SHCkind)
                xmemsaftyManager.append(mem.saftyManager)
                xmemsaftyPromoter.append(mem.saftyPromoter)
                xmemleadEngineer.append(mem.leadEngineer)
                xmemfile0.append(mem.file0)
                xmemInsurance_station.append(mem.Insurance_station)
                xmemInsurance_station_number.append(mem.Insurance_station_number)
                xmemhealthInsurance.append(mem.healthInsurance)
                xmemHIlast4digits.append(mem.HIlast4digits)
                xmemfile1.append(mem.file1)
                xmempensionInsurance.append(mem.pensionInsurance)
                xmemfile2.append(mem.file2)
                xmememploymentinsurance.append(mem.employmentinsurance)
                xmemEIlast4digits.append(mem.EIlast4digits)
                xmemfile3.append(mem.file3)
                xmemretirementCooperation.append(mem.retirementCooperation)
                xmemfile4.append(mem.file4)
                xmemspecialEducation.append(mem.specialEducation)
                xmemfile5.append(mem.file5)
                xmemskillTraning.append(mem.skillTraning)
                xmemfile6.append(mem.file6)
                xmemlicence.append(mem.licence)
                xmemfile7.append(mem.file7)
                xmemdriverLicence_number.append(mem.driverLicence_number)
                xmemfile8.append(mem.file8)
                xmemdateOfAdmission.append(mem.dateOfAdmission)
                xmemfile9.append(mem.file9)
                xmemfile10.append(mem.file10)
                xmemconstractionPermit.append(mem.constractionPermit)
                xmemconstractionPermit_chuji.append(mem.constractionPermit_chuji)
                xmemconstractionPermit_ippan.append(mem.constractionPermit_ippan)
                xmemconstractionPermit_nendo.append(mem.constractionPermit_nendo)
                xmemconstractionPermit_num.append(mem.constractionPermit_num)
                xmemconstractionPermit_gat.append(mem.constractionPermit_gat)
                xmemfile5.append(mem.file5)
                xmemcontractDay.append(mem.contractDay)
                xmemfirstActionDay.append(mem.firstActionDay)
                xmemfinishActionDay.append(mem.finishActionDay)
                xmemcontractionDetail.append(mem.contractionDetail)
                xmemNo_1_specific_skill_foreigner.append(mem.No_1_specific_skill_foreigner)
                xmemforeignWorker.append(mem.foreignWorker)
                xmemforeignIntern.append(mem.foreignIntern)
                xmemcarryonMachine.append(mem.carryonMachine)
                xmemmvehicle.append(mem.mvehicle)
                xmemmobileCrane.append(mem.mobileCrane)
                xmemdangerous.append(mem.dangerous)
            # print(xmemInsurance_station)
        
     
        print(xvehicleModel)
        print(xfirstInspection)
        print(xvehicleNumber)
        print(xfirstUse)
        print(dvName)
    
    #基本職方all **********************************************************************************************
    xxcode =[]
    xxdjgroup =[]
    xxname =[]
    xxkanaName =[]
    xxdjgroup2 =[]
    xxoccupation =[]
    xxrefer =[]
    xxbirthday =[]
    xxage =[]
    xxdateOfEmployment =[]
    xxworkingFirstDate =[]
    xxyearsOfExperience =[]
    xxpostalCode =[]
    xxaddress =[]
    xxtelephoneNumber =[]
    xxfamilyAdd =[]
    xxtel =[]
    xxmedicalCheckUp =[]
    xxbloodPressure =[]
    xxbloodType =[]
    xxspecialHealthCheck =[]
    xxSHCkind =[]
    xxsaftyManager =[]
    xxsaftyPromoter =[]
    xxleadEngineer =[]
    xxfile0 =[]
    xxInsurance_station =[]
    xxInsurance_station_number =[]
    xxhealthInsurance =[]
    xxHIlast4digits =[]
    xxfile1 =[]
    xxpensionInsurance =[]
    xxfile2 =[]
    xxemploymentinsurance =[]
    xxEIlast4digits =[]
    xxfile3 =[]
    xxretirementCooperation =[]
    xxfile4 =[]
    xxfile5 =[] 
    xxfile6 =[]
    xxfile7 =[]
    xxdriverLicence_number =[]
    xxfile8 =[]
    xxdateOfAdmission =[]
    xxfile9 =[]
    xxfile10 =[]
    xxsp=[]
    dxxsp={}
    xxspkey=[]
    xxspvalue=[]
    xxskt=[]
    dxxskt={}
    xxsktkey=[]
    xxsktvalue=[]
    xxlic=[]
    dxxlic={}
    xxlkey=[]
    xxlvalue=[]
    for i in range(len(Xx_all)):
        for x2 in Xx_all:
            for Xx in x2.all():
                xxcode.append(Xx.code)
                xxdjgroup.append(Xx.djgroup)
                xxname.append(Xx.name)
                xxkanaName.append(Xx.kanaName)
                xxdjgroup2.append(Xx.djgroup2)
                xxoccupation.append(Xx.occupation)
                xxrefer.append(Xx.refer)
                xxbirthday.append(Xx.birthday)
                xxage.append(Xx.age)
                xxdateOfEmployment.append(Xx.dateOfEmployment)
                xxworkingFirstDate.append(Xx.workingFirstDate)
                xxyearsOfExperience.append(Xx.yearsOfExperience)
                xxpostalCode.append(Xx.postalCode)
                xxaddress.append(Xx.address)
                xxtelephoneNumber.append(Xx.telephoneNumber)
                xxfamilyAdd.append(Xx.familyAdd)
                xxtel.append(Xx.tel)
                xxmedicalCheckUp.append(Xx.medicalCheckUp)
                xxbloodPressure.append(Xx.bloodPressure)
                xxbloodType.append(Xx.bloodType)
                xxspecialHealthCheck.append(Xx.specialHealthCheck)
                xxSHCkind.append(Xx.SHCkind)
                xxsaftyManager.append(Xx.saftyManager)
                xxsaftyPromoter.append(Xx.saftyPromoter)
                xxleadEngineer.append(Xx.leadEngineer)
                xxfile0.append(Xx.file0)
                xxInsurance_station.append(Xx.Insurance_station)
                xxInsurance_station_number.append(Xx.Insurance_station_number)
                xxhealthInsurance.append(Xx.healthInsurance)
                xxHIlast4digits.append(Xx.HIlast4digits)
                xxfile1.append(Xx.file1)
                xxpensionInsurance.append(Xx.pensionInsurance)
                xxfile2.append(Xx.file2)
                xxemploymentinsurance.append(Xx.employmentinsurance)
                xxEIlast4digits.append(Xx.EIlast4digits)
                xxfile3.append(Xx.file3)
                xxretirementCooperation.append(Xx.retirementCooperation)                    
                xxfile4.append(Xx.file4)
                
                xxdriverLicence_number.append(Xx.driverLicence_number)
                xxfile8.append(Xx.file8)
                xxdateOfAdmission.append(Xx.dateOfAdmission)
                xxfile9.append(Xx.file9)
                xxfile10.append(Xx.file10)
                
                for sp in Xx.specialEducation.all():
                    xxspkey.append(Xx.name)
                    xxspvalue.append(sp.name)
                    for m,n in zip(xxspkey,xxspvalue):
                        l=(m,n)
                        xxsp.append(l)
                        dxxsp=set(xxsp)
                xxfile5.append(Xx.file5)
                
                for skt in Xx.skillTraning.all():
                    xxsktkey.append(Xx.name)
                    xxsktvalue.append(skt.name)
                    for m,n in zip(xxsktkey,xxsktvalue):
                        l=(m,n)
                        xxskt.append(l)
                        dxxskt=set(xxskt)
                xxfile6.append(Xx.file6)
                
                for lic in Xx.licence.all():
                    xxlkey.append(Xx.name)
                    xxlvalue.append(lic.name)
                    for m,n in zip(xxlkey,xxlvalue):
                        l=(m,n)
                        xxlic.append(l)
                        dxxlic=set(xxlic)
                xxfile7.append(Xx.file7)#*********************************************************************         
    
    wb= load_workbook (filename=r"C:\Users\PCUSER\Desktop\tesdata\input_data\グリーンファイル.xlsx",read_only=False,keep_vba=True)
    ws=wb['Sheet1']
    ws.cell(2,5,value=spotName)
    ws.cell(2,6,value=superVisor)
        
    for i in range(0,len(gContractor_all)):
        ws.cell(3+i,5+i,value=gContractor_all[i].gName)
        ws.cell(3+i,6+i,value=gContractor_all[i].postalCode)
        ws.cell(3+i,7+i,value=gContractor_all[i].address)
        ws.cell(3+i,8+i,value=gContractor_all[i].telephoneNumber)
        ws.cell(3+i,9+i,value=gContractor_all[i].FaxNumber)
  
    for i in range(0,len(xrecentName)):
        ws.cell(4+i,5,value=xrecentName[i])
        ws.cell(4+i,6,value=xpostalCode[i])
        ws.cell(4+i,7,value=xaddress[i])
        ws.cell(4+i,8,value=xtelephoneNumber[i])
        ws.cell(4+i,9,value=xFaxNumber[i])
        ws.cell(4+i,10,value=xrepresentative[i])
        ws.cell(4+i,11,value=xsaftyManager[i])
        ws.cell(4+i,12,value=xsaftyPromoter[i])
        ws.cell(4+i,13,value=xleadEngineer[i])
        # ws.cell(4+i,14,value=xfile0[i])
        ws.cell(4+i,15,value=xInsurance_station[i])
        ws.cell(4+i,16,value=xInsurance_station_number[i])
        ws.cell(4+i,17,value=xhealthInsurance[i])
        # ws.cell(4+i,18,value=xfile1[i])
        ws.cell(4+i,19,value=xpensionInsurance[i])
        # ws.cell(4+i,20,value=xfile2[i])
        ws.cell(4+i,21,value=xemploymentinsurance[i])
        ws.cell(4+i,22,value=xemploymentinsurance_number[i])
        # ws.cell(4+i,23,value=xfile3[i])
        ws.cell(4+i,24,value=xretirementCooperation[i])
        # ws.cell(4+i,25,value=xfile4[i])
        ws.cell(4+i,26,value=xconstractionPermit[i])
        ws.cell(4+i,27,value=xconstractionPermit_chuji[i])
        ws.cell(4+i,28,value=xconstractionPermit_ippan[i])
        ws.cell(4+i,29,value=xconstractionPermit_nendo[i])
        ws.cell(4+i,30,value=xconstractionPermit_num[i])
        ws.cell(4+i,31,value=xconstractionPermit_gat[i])
        # ws.cell(4+i,32,value=xfile5[i])
        ws.cell(4+i,33,value=xcontractDay[i])
        ws.cell(4+i,34,value=xfirstActionDay[i])
        ws.cell(4+i,35,value=xfinishActionDay[i])
        ws.cell(4+i,36,value=xcontractionDetail[i])
        ws.cell(4+i,37,value=xNo_1_specific_skill_foreigner[i])
        ws.cell(4+i,38,value=xforeignWorker[i])
        ws.cell(4+i,39,value=xforeignIntern[i])
        ws.cell(4+i,40,value=xcarryonMachine[i])
        ws.cell(4+i,41,value=xvehicleNumber[i])
        ws.cell(4+i,42,value=xmrecent[i])
        ws.cell(4+i,43,value=xvehicleNumber[i])
        ws.cell(4+i,44,value=xfirstUse[i])
        ws.cell(4+i,45,value=xfinishUse[i])
        ws.cell(4+i,46,value=xvehicleModel[i])
        ws.cell(4+i,47,value=xfirstInspection[i])
        ws.cell(4+i,48,value=xfinishInspection[i])
        # ws.cell(4+i,49,value=xfile1[i])
        ws.cell(4+i,50,value=xliabilityComp[i])
        ws.cell(4+i,51,value=xliability_num[i])
        ws.cell(4+i,52,value=xfirstliability[i])
        ws.cell(4+i,53,value=xfinishliability[i])
        # ws.cell(4+i,54,value=xfile2[i])
        ws.cell(4+i,55,value=xvoluntaryInsureComp[i])
        ws.cell(4+i,56,value=xvoluntaryInsureNum[i])
        ws.cell(4+i,57,value=xfirstVoluntary[i])
        ws.cell(4+i,58,value=xfinishVoluntary[i])
        ws.cell(4+i,59,value=xvoluntary_object[i])
        ws.cell(4+i,60,value=xvoluntary_human[i])
        ws.cell(4+i,61,value=xvoluntary_passenger[i])
        # ws.cell(4+i,62,value=xfile3[i])
        ws.cell(4+i,63,value=dvName[i])
        ws.cell(4+i,64,value=xdrive_departure[i])
        ws.cell(4+i,65,value=xdrive_waypoint1[i])
        ws.cell(4+i,66,value=xdrive_waypoint2[i])
        ws.cell(4+i,67,value=xdrive_arrival[i])
        ws.cell(4+i,68,value=xmobileCrane[i])
        ws.cell(4+i,69,value=xdangerous[i])
        ws.cell(4+i,70,value=xmemcode[i])
        ws.cell(4+i,71,value=xmemdjgroup[i])
        ws.cell(4+i,72,value=xmemname[i])
        ws.cell(4+i,73,value=xmemkanaName[i])
        ws.cell(4+i,74,value=xmemdjgroup2[i])
        ws.cell(4+i,75,value=xmemoccupation[i])
        ws.cell(4+i,76,value=xmemrefer[i])
        ws.cell(4+i,77,value=xmembirthday[i])
        ws.cell(4+i,78,value=xmemage[i])
        ws.cell(4+i,79,value=xmemdateOfEmployment[i])
        ws.cell(4+i,80,value=xmemworkingFirstDate[i])
        ws.cell(4+i,81,value=xmemyearsOfExperience[i])
        ws.cell(4+i,82,value=xmempostalCode[i])
        ws.cell(4+i,83,value=xmemaddress[i])
        ws.cell(4+i,84,value=xmemtelephoneNumber[i])
        ws.cell(4+i,85,value=xmemfamilyAdd[i])
        ws.cell(4+i,86,value=xmemtel[i])
        ws.cell(4+i,87,value=xmemmedicalCheckUp[i])
        ws.cell(4+i,88,value=xmembloodPressure[i])
        ws.cell(4+i,89,value=xmembloodType[i])
        ws.cell(4+i,90,value=xmemspecialHealthCheck[i])
        ws.cell(4+i,91,value=xmemSHCkind[i])
        ws.cell(4+i,92,value=xmemsaftyManager[i])
        ws.cell(4+i,93,value=xmemsaftyPromoter[i])
        ws.cell(4+i,94,value=xmemleadEngineer[i])
        # ws.cell(4+i,95,value=xmemfile0[i])
        ws.cell(4+i,96,value=xmemInsurance_station[i])
        ws.cell(4+i,97,value=xmemInsurance_station_number[i])
        ws.cell(4+i,98,value=xmemhealthInsurance[i])
        ws.cell(4+i,99,value=xmemHIlast4digits[i])
        # ws.cell(4+i,100,value=xmemfile1[i])
        ws.cell(4+i,101,value=xmempensionInsurance[i])
        # ws.cell(4+i,102,value=xmemfile2[i])
        ws.cell(4+i,103,value=xmememploymentinsurance[i])
        ws.cell(4+i,104,value=xmemEIlast4digits[i])
        # ws.cell(4+i,105,value=xmemfile3[i])
        ws.cell(4+i,106,value=xmemretirementCooperation[i])
        # ws.cell(4+i,107,value=xmemfile4[i])
        # ws.cell(4+i,108,value=xmemspecialEducation[i])
        # ws.cell(4+i,109,value=xmemfile6[i])
        # ws.cell(4+i,110,value=xmemlicence.name[i])
        # ws.cell(4+i,111,value=xmemfile7[i])
        # ws.cell(4+i,112,value=xmemdriverLicence_number[i])
        # ws.cell(4+i,113,value=xmemfile8[i])
        ws.cell(4+i,114,value=xmemdateOfAdmission[i])
        # ws.cell(4+i,115,value=xmemfile9[i])
        # ws.cell(4+i,116,value=xmemfile10[i])
        ws.cell(4+i,117,value=xmemconstractionPermit[i])
        ws.cell(4+i,118,value=xmemconstractionPermit_chuji[i])
        ws.cell(4+i,119,value=xmemconstractionPermit_ippan[i])
        ws.cell(4+i,120,value=xmemconstractionPermit_nendo[i])
        ws.cell(4+i,121,value=xmemconstractionPermit_num[i])
        ws.cell(4+i,122,value=xmemconstractionPermit_gat[i])
        # ws.cell(4+i,123,value=xmemfile5[i])
        ws.cell(4+i,124,value=xmemcontractDay[i])
        ws.cell(4+i,125,value=xmemfirstActionDay[i])
        ws.cell(4+i,126,value=xmemfinishActionDay[i])
        ws.cell(4+i,127,value=xmemcontractionDetail[i])
        ws.cell(4+i,128,value=xmemNo_1_specific_skill_foreigner[i])
        ws.cell(4+i,129,value=xmemforeignWorker[i])
        ws.cell(4+i,130,value=xmemforeignIntern[i])
        ws.cell(4+i,131,value=xmemcarryonMachine[i])
        # ws.cell(4+i,132,value=xmemmvehicle[i])
        ws.cell(4+i,133,value=xmemmobileCrane[i])
        ws.cell(4+i,134,value=xmemdangerous[i])
        
    for i in range(len(xxname)):
        ws.cell(6+i,5,value=xxname[i]) 
        ws.cell(6+i,6,value=xxcode [i])
        ws.cell(6+i,7,value=xxdjgroup [i])
        ws.cell(6+i,8,value=xxname [i])
        ws.cell(6+i,9,value=xxkanaName [i])
        ws.cell(6+i,10,value=xxdjgroup2 [i])
        ws.cell(6+i,11,value=xxoccupation [i])
        ws.cell(6+i,12,value=xxrefer [i])
        ws.cell(6+i,13,value=xxbirthday [i])
        ws.cell(6+i,14,value=xxage [i])
        ws.cell(6+i,15,value=xxdateOfEmployment [i])
        ws.cell(6+i,16,value=xxworkingFirstDate [i])
        ws.cell(6+i,17,value=xxyearsOfExperience [i])
        ws.cell(6+i,18,value=xxpostalCode [i])
        ws.cell(6+i,19,value=xxaddress [i])
        ws.cell(6+i,20,value=xxtelephoneNumber [i])
        ws.cell(6+i,21,value=xxfamilyAdd [i])
        ws.cell(6+i,22,value=xxtel [i])
        ws.cell(6+i,23,value=xxmedicalCheckUp [i])
        ws.cell(6+i,24,value=xxbloodPressure [i])
        ws.cell(6+i,25,value=xxbloodType [i])
        ws.cell(6+i,26,value=xxspecialHealthCheck [i])
        ws.cell(6+i,27,value=xxSHCkind [i])
        ws.cell(6+i,28,value=xxsaftyManager [i])
        ws.cell(6+i,29,value=xxsaftyPromoter [i])
        ws.cell(6+i,30,value=xxleadEngineer [i])
        # ws.cell(6+i,31,value=xxfile0 [i])
        ws.cell(6+i,32,value=xxInsurance_station [i])
        ws.cell(6+i,33,value=xxInsurance_station_number [i])
        ws.cell(6+i,34,value=xxhealthInsurance [i])
        ws.cell(6+i,35,value=xxHIlast4digits [i])
        # ws.cell(6+i,36,value=xxfile1 [i])
        ws.cell(6+i,37,value=xxpensionInsurance [i])
        # ws.cell(6+i,38,value=xxfile2 [i])
        ws.cell(6+i,39,value=xxemploymentinsurance [i])
        ws.cell(6+i,40,value=xxEIlast4digits [i])
        # ws.cell(6+i,41,value=xxfile3 [i])
        ws.cell(6+i,42,value=xxretirementCooperation [i])
        # ws.cell(6+i,43,value=xxfile4 [i])
        # ws.cell(6+i,44,value=xxfile5 [i])
    
        # ws.cell(6+i,45,value=xxfile6 [i])
    
        # ws.cell(6+i,46,value=xxfile7 [i])
        ws.cell(6+i,47,value=xxdriverLicence_number [i])
        # ws.cell(6+i,48,value=xxfile8 [i])
        ws.cell(6+i,49,value=xxdateOfAdmission [i])
        # ws.cell(6+i,50,value=xxfile9 [i])
        # ws.cell(6+i,51,value=xxfile10 [i])
        
    for m,i in zip(dxxsp,range(len(dxxsp))):
            ws.cell(30+i,5,value=m[0]+':'+m[1])
            
    for m,i in zip(dxxskt,range(len(dxxskt))):
            ws.cell(45+i,5,value=m[0]+':'+m[1])
 
    for m,i in zip(dxxlic,range(len(dxxlic))): 
            ws.cell(60+i,5,value=m[0]+':'+m[1])
            
    
    wb.save(r"C:\Users\PCUSER\Desktop\tesdata\output_data\{0:%Y%m%d_%H%M%S}.xlsm".format(now))
    print('finished')
    
if __name__=="__main__":
    func()
    